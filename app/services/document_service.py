"""
services/document_service.py
-----------------------------
Service d'extraction de texte et résumé IA pour les documents uploadés.

Supporte : PDF, Word (.docx), Excel (.xlsx), images (nom seulement).
Tronque le texte extrait à ~10000 caractères pour les prompts IA.
"""

import logging
import os

from django.conf import settings

from app.services.scoring_llm import _get_client

logger = logging.getLogger(__name__)

MAX_TEXTE_EXTRAIT = 10000
MODEL_RESUME = "mistral-small-latest"


def extraire_texte(fichier_path):
    """
    Extrait le texte d'un fichier selon son extension.

    Args:
        fichier_path: Chemin absolu du fichier.

    Returns:
        str: Texte extrait (tronqué à MAX_TEXTE_EXTRAIT chars).
    """
    ext = os.path.splitext(fichier_path)[1].lower()

    try:
        if ext == '.pdf':
            return _extraire_pdf(fichier_path)
        elif ext == '.docx':
            return _extraire_docx(fichier_path)
        elif ext in ('.xlsx', '.xls'):
            return _extraire_xlsx(fichier_path)
        elif ext in ('.png', '.jpg', '.jpeg', '.gif', '.webp'):
            return f"[Image : {os.path.basename(fichier_path)}]"
        elif ext in ('.txt', '.csv', '.md'):
            with open(fichier_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()[:MAX_TEXTE_EXTRAIT]
        else:
            return f"[Document : {os.path.basename(fichier_path)} — format non supporté pour l'extraction]"
    except Exception as e:
        logger.error("Erreur extraction texte %s : %s", fichier_path, e)
        return f"[Erreur extraction : {os.path.basename(fichier_path)}]"


def _extraire_pdf(path):
    """Extrait le texte d'un PDF via PyPDF2."""
    from PyPDF2 import PdfReader
    reader = PdfReader(path)
    texte = []
    for page in reader.pages:
        t = page.extract_text()
        if t:
            texte.append(t)
        if len("\n".join(texte)) > MAX_TEXTE_EXTRAIT:
            break
    return "\n".join(texte)[:MAX_TEXTE_EXTRAIT]


def _extraire_docx(path):
    """Extrait le texte d'un fichier Word .docx."""
    from docx import Document
    doc = Document(path)
    texte = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    return texte[:MAX_TEXTE_EXTRAIT]


def _extraire_xlsx(path):
    """Extrait le contenu d'un fichier Excel."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    texte = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        texte.append(f"--- Feuille : {sheet} ---")
        for row in ws.iter_rows(max_row=200, values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            line = " | ".join(cells).strip()
            if line and line != "|":
                texte.append(line)
            if len("\n".join(texte)) > MAX_TEXTE_EXTRAIT:
                break
    wb.close()
    return "\n".join(texte)[:MAX_TEXTE_EXTRAIT]


def generer_resume_document(document_id):
    """
    Génère un résumé IA du document extrait via Mistral.

    Args:
        document_id: ID du DocumentTitre.

    Returns:
        str: Résumé ou None si erreur.
    """
    from app.models import DocumentTitre

    try:
        doc = DocumentTitre.objects.select_related('titre').get(pk=document_id)
    except DocumentTitre.DoesNotExist:
        return None

    if not doc.texte_extrait or doc.texte_extrait.startswith('['):
        return None

    # Tronquer pour le prompt
    texte = doc.texte_extrait[:6000]

    prompt = f"""Résume ce document lié à {doc.titre.nom} ({doc.titre.ticker}), secteur {doc.titre.secteur or 'inconnu'}.
Type de document : {doc.get_type_doc_display()}
Nom : {doc.nom}

CONTENU :
{texte}

Rédige un résumé en 3-5 phrases en français, en langage SIMPLE pour un débutant en bourse.
Mets en avant les points clés qui impactent la valeur du titre (chiffres, résultats, perspectives).
Si c'est une étude clinique, résume le résultat et ce que ça signifie pour l'entreprise.
Réponds directement sans titre."""

    try:
        client = _get_client()
        response = client.chat.complete(
            model=MODEL_RESUME,
            messages=[
                {"role": "system", "content": "Tu résumes des documents financiers en langage simple pour un débutant."},
                {"role": "user", "content": prompt},
            ],
            max_tokens=300,
            temperature=0.2,
        )
        resume = response.choices[0].message.content.strip()

        doc.resume_ia = resume
        doc.save(update_fields=['resume_ia'])

        logger.info("[Doc] Résumé généré pour document %d (%s)", document_id, doc.nom)
        return resume

    except Exception as e:
        logger.error("[Doc] Erreur résumé document %d : %s", document_id, e)
        return None


def traiter_document(document_id):
    """
    Pipeline complet : extraction texte + résumé IA.
    Appelé après l'upload d'un document.
    """
    from app.models import DocumentTitre

    try:
        doc = DocumentTitre.objects.get(pk=document_id)
    except DocumentTitre.DoesNotExist:
        return

    # 1. Extraction texte
    fichier_path = doc.fichier.path
    texte = extraire_texte(fichier_path)
    doc.texte_extrait = texte
    doc.taille = os.path.getsize(fichier_path)
    doc.save(update_fields=['texte_extrait', 'taille'])

    # 2. Résumé IA
    generer_resume_document(document_id)
